from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import io
from openpyxl import Workbook
from openpyxl.styles import Font

from ..db.session import get_db
from ..models import Internship, Company, User, Application, Resume
from ..utils.security import get_current_user
from ..schemas import InternshipCreate, InternshipOut, ApplicationOut  # Reuse InternshipCreate for job creation
from ..core.config import settings

router = APIRouter(prefix="/hr/job", tags=["hr-jobs"])

def require_hr_role(current_user: User = Depends(get_current_user)):
    """Dependency to ensure user is HR"""
    if current_user.role != "hr":
        raise HTTPException(status_code=403, detail="Only HR users can access this endpoint")
    return current_user

@router.post("/create", response_model=InternshipOut)
def create_job(
    job: InternshipCreate,
    current_user: User = Depends(require_hr_role),
    db: Session = Depends(get_db)
):
    """
    Create a new internship job posting.
    Requires HR role.
    """
    # Find or create company
    company = db.query(Company).filter(Company.name == job.company_name).first()
    if not company:
        company = Company(name=job.company_name)
        db.add(company)
        db.commit()
        db.refresh(company)

    # Create internship record
    db_internship = Internship(
        title=job.title,
        company_id=company.id,
        location=job.location,
        stipend=getattr(job, 'stipend', None),  # Optional stipend
        description=job.description,
        min_qualifications=job.min_qualifications,
        expected_qualifications=job.expected_qualifications,
        deadline=job.deadline,
        posted_by=current_user.id
    )
    db.add(db_internship)
    db.commit()
    db.refresh(db_internship)

    return InternshipOut(
        id=db_internship.id,
        title=db_internship.title,
        company_name=company.name,
        location=db_internship.location,
        stipend=db_internship.stipend,
        description=db_internship.description,
        min_qualifications=db_internship.min_qualifications,
        expected_qualifications=db_internship.expected_qualifications,
        deadline=db_internship.deadline,
        posted_at=db_internship.posted_at,
        posted_by_name=current_user.name
    )

@router.get("/applicants", response_model=List[dict])
def get_job_applicants(
    job_id: int = Query(..., description="ID of the job/internship"),
    current_user: User = Depends(require_hr_role),
    db: Session = Depends(get_db)
):
    """
    Get list of applicants for a specific job.
    Includes applicant name, resume path, and status.
    Requires HR role.
    """
    # Verify the job exists and was posted by current HR or belongs to their company (additional security)
    internship = db.query(Internship).filter(Internship.id == job_id, Internship.posted_by == current_user.id).first()
    if not internship:
        raise HTTPException(status_code=404, detail="Job not found or you don't have access")

    applications = db.query(Application).join(User).join(Resume).filter(
        Application.internship_id == job_id
    ).all()

    result = []
    for app in applications:
        result.append({
            "applicant_name": app.user.name,
            "resume_path": app.resume.file_path,
            "status": app.status,
            "applied_at": app.applied_at
        })
    return result

@router.get("/export")
def export_job_applicants(
    job_id: int = Query(..., description="ID of the job/internship"),
    current_user: User = Depends(require_hr_role),
    db: Session = Depends(get_db)
):
    """
    Export applicants for a job as XLSX file.
    Columns: Candidate Name, Role, Resume Path, Status
    Requires HR role.
    """
    # Verify access (same as above)
    internship = db.query(Internship).filter(Internship.id == job_id, Internship.posted_by == current_user.id).first()
    if not internship:
        raise HTTPException(status_code=404, detail="Job not found or you don't have access")

    applications = db.query(Application).join(User).join(Resume).filter(
        Application.internship_id == job_id
    ).all()

    # Create workbook in memory
    wb = Workbook()
    ws = wb.active
    ws.title = "Applicants"

    # Headers
    headers = ["Candidate Name", "Role", "Resume Path", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Data rows
    for row, app in enumerate(applications, 2):
        ws.cell(row=row, column=1, value=app.user.name)
        ws.cell(row=row, column=2, value=app.user.role)
        ws.cell(row=row, column=3, value=app.resume.file_path)
        ws.cell(row=row, column=4, value=app.status)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"applicants_job_{job_id}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Security Comments:
# 1. Role-based access: Already enforced with require_hr_role dependency. Consider adding company ownership check for multi-company HR.
# 2. Input validation: Pydantic schemas handle this; add custom validators for deadline > now, etc.
# 3. Rate limiting: Use slowapi or similar middleware to prevent abuse on export endpoints.
# 4. File security: For exports, ensure sensitive data (e.g., full resumes) isn't exposed; here only paths are included.
# 5. Authentication: All endpoints use JWT via get_current_user; ensure tokens expire and refresh properly.
# 6. SQL Injection: SQLAlchemy prevents this; use parameterized queries.
# 7. Auditing: Log job creations and exports for compliance.
# 8. HTTPS: Enforce in production to protect auth tokens.
