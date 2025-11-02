from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from typing import Optional, List
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from ..database import get_db
from ..models.application import Application
from ..models.internship import Internship
from ..models.user import User
from ..models.resume import Resume
from ..models.company import Company
from ..dependencies import get_current_hr_user as get_current_hr
from ..schemas.application import ApplicationOut
from ..schemas.internship import InternshipOut

class ApplicationStatusUpdate(BaseModel):
    status: str  # 'pending', 'accepted', 'rejected'

router = APIRouter()

@router.get("/my_jobs", response_model=List[InternshipOut])
def get_my_jobs(
    current_hr: User = Depends(get_current_hr),
    db: Session = Depends(get_db)
):
    """
    Get all jobs/internships posted by the current HR user.
    """
    internships = db.query(Internship).filter(Internship.posted_by == current_hr.id).all()
    result = []
    for i in internships:
        result.append(InternshipOut(
            id=i.id,
            title=i.title,
            company_name=i.company.name,
            location=i.location,
            stipend=i.stipend,
            description=i.description,
            min_qualifications=i.min_qualifications,
            expected_qualifications=i.expected_qualifications,
            deadline=i.deadline,
            posted_at=i.posted_at,
            posted_by_name=i.posted_by_user.name
        ))
    return result

@router.get("/my_jobs/{hr_id}/applications", response_model=List[ApplicationOut])
def get_hr_job_applications(
    hr_id: int,
    current_hr: User = Depends(get_current_hr),
    db: Session = Depends(get_db)
):
    """
    Get applications for internships posted by the specified HR user.
    Only accessible if the requesting user is the same HR or an admin.
    """
    if hr_id != current_hr.id:
        raise HTTPException(status_code=403, detail="Can only access your own job applications")

    # Get internships posted by this HR
    internships = db.query(Internship).filter(Internship.posted_by == hr_id).all()
    internship_ids = [i.id for i in internships]

    # Get applications for these internships
    applications = db.query(Application).join(Internship).join(Resume).join(User, Application.user_id == User.id).join(Company, Internship.company_id == Company.id).filter(Application.internship_id.in_(internship_ids)).all()

    result = []
    for a in applications:
        result.append(ApplicationOut(
            id=a.id,
            internship_id=a.internship.id,
            internship_title=a.internship.title,
            internship_company=a.internship.company.name,
            resume_id=a.resume.id,
            resume_title=a.resume.title,
            cover_letter=a.cover_letter,
            status=a.status,
            applied_at=a.applied_at
        ))
    return result

def _get_job_applicants_logic(job_id: int, current_hr: User, db: Session):
    """Shared logic for getting job applicants"""
    # Verify job belongs to HR
    internship = db.query(Internship).filter(Internship.id == job_id, Internship.posted_by == current_hr.id).first()
    if not internship:
        raise HTTPException(status_code=404, detail="Job not found or not authorized")

    # Get applications for this job
    applications = db.query(Application).join(Resume).join(User, Application.user_id == User.id).filter(Application.internship_id == job_id).all()

    result = []
    for a in applications:
        result.append({
            'id': a.id,
            'applicant_name': a.user.name,
            'resume_id': a.resume.id,
            'resume_title': a.resume.title,
            'resume_path': a.resume.file_path,
            'status': a.status,
            'applied_at': a.applied_at
        })
    return result

@router.get("/job/{job_id}/applicants")
def get_job_applicants(
    job_id: int,
    current_hr: User = Depends(get_current_hr),
    db: Session = Depends(get_db)
):
    """Get applicants for a specific job posted by the HR."""
    return _get_job_applicants_logic(job_id, current_hr, db)

@router.get("/applications/{job_id}")
def get_applications_for_job(
    job_id: int,
    current_hr: User = Depends(get_current_hr),
    db: Session = Depends(get_db)
):
    """Alias endpoint for frontend compatibility"""
    return _get_job_applicants_logic(job_id, current_hr, db)

@router.put("/applications/{application_id}/status")
def update_application_status(
    application_id: int,
    status_update: ApplicationStatusUpdate,
    current_hr: User = Depends(get_current_hr),
    db: Session = Depends(get_db)
):
    """
    Update the status of an application.
    HR can only update applications for their own job postings.
    """
    # Get the application
    application = db.query(Application).filter(Application.id == application_id).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    
    # Verify the internship belongs to the HR user
    internship = db.query(Internship).filter(
        Internship.id == application.internship_id,
        Internship.posted_by == current_hr.id
    ).first()
    if not internship:
        raise HTTPException(status_code=403, detail="Not authorized to update this application")
    
    # Validate status
    valid_statuses = ['pending', 'accepted', 'rejected']
    if status_update.status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}")
    
    # Update status
    application.status = status_update.status
    db.commit()
    db.refresh(application)
    
    return {"message": "Application status updated successfully", "status": application.status}

@router.get("/export/{job_id}")
def export_job_applications(
    job_id: int,
    current_hr: User = Depends(get_current_hr),
    db: Session = Depends(get_db)
):
    """
    Export applications for a specific job as XLSX file.
    """
    # Verify the job belongs to the HR
    internship = db.query(Internship).filter(Internship.id == job_id, Internship.posted_by == current_hr.id).first()
    if not internship:
        raise HTTPException(status_code=404, detail="Job not found or not authorized")

    # Get applications for this job
    applications = db.query(Application).join(Resume).join(User, Application.user_id == User.id).filter(Application.internship_id == job_id).all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Applications"

    # Headers
    headers = ["Candidate Name", "Role", "Resume Path", "Status", "Applied At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data
    for row, app in enumerate(applications, 2):
        ws.cell(row=row, column=1, value=app.user.name)
        ws.cell(row=row, column=2, value=app.user.role)
        ws.cell(row=row, column=3, value=app.resume.file_path)
        ws.cell(row=row, column=4, value=app.status)
        ws.cell(row=row, column=5, value=app.applied_at)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"job_{job_id}_applications.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/post_job")
def post_hr_job(
    internship_data: dict,
    current_hr: User = Depends(get_current_hr),
    db: Session = Depends(get_db)
):
    """
    Post a new job/internship (wrapper for /internships/, HR only).
    """
    # Forward to internships router logic, but ensure HR role (already checked)
    from .internships import create_internship
    # Note: This is a simplified forward; in practice, use shared service or include the router
    # For now, replicate the logic or call the function if possible
    # Assuming we can import and call
    return create_internship(internship_data, current_hr, db)
